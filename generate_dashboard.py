import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.drawing.image import Image
import io

# --- Nomes Exatos das Abas ---
BASES_SHEET_NAME = 'B̳ases'
CALCULOS_SHEET_NAME = 'C̳álculos'
DASHBOARD_SHEET_NAME = 'D̳ashboard'

# --- 1. Carregar os Dados ---
try:
    # Carregar o DataFrame usando o nome exato da aba
    df = pd.read_excel('base.xlsx', sheet_name=BASES_SHEET_NAME)
except Exception as e:
    print(f"Erro ao carregar a aba '{BASES_SHEET_NAME}': {e}")
    exit()
except FileNotFoundError:
    print("Erro: O arquivo 'base.xlsx' não foi encontrado.")
    exit()

# Renomear colunas com quebra de linha
df.columns = [col.replace('\n', ' ') for col in df.columns]

# --- 2. Pré-processamento e Limpeza de Dados ---
# Garantir que as colunas de valor sejam numéricas
df['Subscription Price'] = pd.to_numeric(df['Subscription Price'], errors='coerce')
# Substituir '-' por 0 antes de converter para numérico
df['EA Play Season Pass Price'] = df['EA Play Season Pass Price'].replace('-', 0)
df['EA Play Season Pass Price'] = pd.to_numeric(df['EA Play Season Pass Price'], errors='coerce')
df['Minecraft Season Pass Price'] = pd.to_numeric(df['Minecraft Season Pass Price'], errors='coerce')
df['Coupon Value'] = pd.to_numeric(df['Coupon Value'], errors='coerce')
df['Total Value'] = pd.to_numeric(df['Total Value'], errors='coerce')

# Preencher valores NaN após coerção com 0
df = df.fillna(0)

# --- 3. Cálculos e Métricas (Sheet 'Cálculos') ---

# Pergunta 1: Faturamento Total de vendas de planos anuais
df_anual = df[df['Subscription Type'] == 'Annual']
faturamento_anual_total = df_anual['Total Value'].sum()

# Pergunta 2: Faturamento Total de vendas de planos anuais, separado por auto renovação
faturamento_anual_auto_renovacao = df_anual.groupby('Auto Renewal')['Total Value'].sum().reset_index()
faturamento_anual_auto_renovacao.columns = ['Auto Renewal', 'Total Value']

# Pergunta 3: Total de Vendas de Assinaturas do EA Play (Preço total)
faturamento_ea_play = df['EA Play Season Pass Price'].sum()
faturamento_ea_play_por_plano = df.groupby('Plan')['EA Play Season Pass Price'].sum().reset_index()
faturamento_ea_play_por_plano.columns = ['Plan', 'EA Play Season Pass Price']

# Pergunta 4: Total de Vendas de Assinaturas do Minecraft Season Pass (Preço total)
faturamento_minecraft = df['Minecraft Season Pass Price'].sum()
faturamento_minecraft_por_plano = df.groupby('Plan')['Minecraft Season Pass Price'].sum().reset_index()
faturamento_minecraft_por_plano.columns = ['Plan', 'Minecraft Season Pass Price']

# Métrica Adicional: Total de Assinantes
total_assinantes = df['Subscriber ID'].nunique()

# Métrica Adicional: Receita Média por Assinante (ARPU)
arpu = df['Total Value'].sum() / total_assinantes

# Métrica Adicional: Distribuição de Planos
distribuicao_planos = df.groupby('Plan')['Subscriber ID'].nunique().reset_index()
distribuicao_planos.columns = ['Plan', 'Count']

# --- 4. Criação do Arquivo Excel Final ---

# Carregar o arquivo base para manter as abas e formatação existentes (Assets, Bases)
wb = load_workbook('base.xlsx')

# --- 4.1. Configurar a aba 'Cálculos' ---
ws_calc = wb[CALCULOS_SHEET_NAME]

# Limpar a aba Cálculos
ws_calc.delete_rows(1, ws_calc.max_row)

# Função auxiliar para escrever DataFrames na planilha
def write_df_to_sheet(ws, df_to_write, start_row, title):
    ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
    start_row += 1
    
    # Escrever cabeçalho
    for c_idx, col_name in enumerate(df_to_write.columns):
        ws.cell(row=start_row, column=c_idx + 1, value=col_name).font = Font(bold=True)
    start_row += 1
    
    # Escrever dados
    for r_idx, row in df_to_write.iterrows():
        for c_idx, value in enumerate(row):
            cell = ws.cell(row=start_row + r_idx, column=c_idx + 1, value=value)
            if 'Value' in df_to_write.columns[c_idx] or 'Price' in df_to_write.columns[c_idx]:
                cell.number_format = '"R$"#,##0.00'
    return start_row + len(df_to_write) + 2

current_row = 1
# Pergunta 1
ws_calc.cell(row=current_row, column=1, value="Faturamento Anual Total").font = Font(bold=True)
ws_calc.cell(row=current_row + 1, column=1, value="Total Value").font = Font(bold=True)
ws_calc.cell(row=current_row + 1, column=2, value=faturamento_anual_total).number_format = '"R$"#,##0.00'
current_row += 3

# Pergunta 2
current_row = write_df_to_sheet(ws_calc, faturamento_anual_auto_renovacao, current_row, "Faturamento Anual por Auto Renovação")
ws_calc.cell(row=current_row - 1, column=3, value="Total Geral").font = Font(bold=True)
ws_calc.cell(row=current_row - 1, column=4, value=faturamento_anual_total).number_format = '"R$"#,##0.00'

# Pergunta 3
current_row = write_df_to_sheet(ws_calc, faturamento_ea_play_por_plano, current_row, "Faturamento EA Play por Plano")
ws_calc.cell(row=current_row - 1, column=3, value="Total Geral").font = Font(bold=True)
ws_calc.cell(row=current_row - 1, column=4, value=faturamento_ea_play).number_format = '"R$"#,##0.00'

# Pergunta 4
current_row = write_df_to_sheet(ws_calc, faturamento_minecraft_por_plano, current_row, "Faturamento Minecraft por Plano")
ws_calc.cell(row=current_row - 1, column=3, value="Total Geral").font = Font(bold=True)
ws_calc.cell(row=current_row - 1, column=4, value=faturamento_minecraft).number_format = '"R$"#,##0.00'

# Distribuição de Planos
current_row = write_df_to_sheet(ws_calc, distribuicao_planos, current_row, "Distribuição de Assinantes por Plano")

# Métricas Chave
ws_calc.cell(row=current_row, column=1, value="Métricas Chave").font = Font(bold=True)
ws_calc.cell(row=current_row + 1, column=1, value="Total de Assinantes").font = Font(bold=True)
ws_calc.cell(row=current_row + 1, column=2, value=total_assinantes).number_format = '#,##0'
ws_calc.cell(row=current_row + 2, column=1, value="Receita Média por Assinante (ARPU)").font = Font(bold=True)
ws_calc.cell(row=current_row + 2, column=2, value=arpu).number_format = '"R$"#,##0.00'

# Ajustar largura das colunas
for col in ws_calc.columns:
    max_length = 0
    column = col[0].column_letter # Get the column letter
    for cell in col:
        try: # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_calc.column_dimensions[column].width = adjusted_width

# --- 4.2. Configurar a aba 'Dashboard' ---
ws_dash = wb[DASHBOARD_SHEET_NAME]

# Limpar a aba Dashboard
ws_dash.delete_rows(1, ws_dash.max_row)

# Configurações de Estilo
fill_header = PatternFill(start_color="9BC848", end_color="9BC848", fill_type="solid")
fill_card = PatternFill(start_color="E8E6E9", end_color="E8E6E9", fill_type="solid")
font_header = Font(bold=True, color="FFFFFF", size=16)
font_metric = Font(bold=True, size=24)
font_label = Font(bold=True, size=12)
border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

# Título do Dashboard
ws_dash.merge_cells('A1:F2')
title_cell = ws_dash['A1']
title_cell.value = "XBOX GAME PASS SUBSCRIPTIONS SALES"
title_cell.fill = fill_header
title_cell.font = font_header
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# --- Cards de Métricas Chave ---
metrics = [
    ("Total de Assinantes", total_assinantes, '#,##0', 'A4'),
    ("Faturamento Anual Total", faturamento_anual_total, '"R$"#,##0.00', 'C4'),
    ("Receita Média por Assinante (ARPU)", arpu, '"R$"#,##0.00', 'E4')
]

for i, (label, value, num_format, start_cell) in enumerate(metrics):
    col_start = ws_dash[start_cell].column
    row_start = ws_dash[start_cell].row
    
    # Mesclar células para o fundo do card (3 linhas x 2 colunas)
    ws_dash.merge_cells(start_row=row_start, start_column=col_start, end_row=row_start + 2, end_column=col_start + 1)
    
    # A única célula que pode ser escrita é a superior esquerda (A4, C4, E4)
    card_cell = ws_dash.cell(row=row_start, column=col_start)
    
    # Aplicar estilo de fundo e borda
    card_cell.fill = fill_card
    card_cell.border = border_thin
    
    # Escrever o Label e o Valor na célula superior esquerda, usando quebra de linha
    # para simular a separação visual.
    # O openpyxl não permite formatar partes do texto, então o estilo será aplicado a todo o conteúdo.
    card_cell.value = f"{label}\n\n{value}"
    card_cell.number_format = num_format
    card_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Ajustar largura das colunas para os cards
ws_dash.column_dimensions['A'].width = 15
ws_dash.column_dimensions['B'].width = 1
ws_dash.column_dimensions['C'].width = 15
ws_dash.column_dimensions['D'].width = 1
ws_dash.column_dimensions['E'].width = 15
ws_dash.column_dimensions['F'].width = 1

# --- Gráfico 1: Faturamento Anual por Auto Renovação ---
chart1 = BarChart()
chart1.type = "col"
chart1.style = 10
chart1.title = "Faturamento Anual por Auto Renovação"
chart1.y_axis.title = "Valor (R$)"
chart1.x_axis.title = "Auto Renovação"

# Dados para o gráfico (Sheet Cálculos)
# O DataFrame faturamento_anual_auto_renovacao começa na linha 7 (5 + 2) e vai até a linha 8.
start_row_auto_renovacao = 7
end_row_auto_renovacao = 8

data1 = Reference(ws_calc, min_col=2, min_row=start_row_auto_renovacao - 1, max_row=end_row_auto_renovacao, max_col=2) # -1 para incluir o cabeçalho
cats1 = Reference(ws_calc, min_col=1, min_row=start_row_auto_renovacao, max_row=end_row_auto_renovacao)

chart1.add_data(data1, titles_from_data=True)
chart1.set_categories(cats1)
chart1.height = 10 # Altura em cm
chart1.width = 15 # Largura em cm
ws_dash.add_chart(chart1, "A8")

# --- Gráfico 2: Distribuição de Assinantes por Plano ---
chart2 = BarChart()
chart2.type = "col"
chart2.style = 10
chart2.title = "Distribuição de Assinantes por Plano"
chart2.y_axis.title = "Contagem"
chart2.x_axis.title = "Plano"

# Dados para o gráfico (Sheet Cálculos)
# O DataFrame distribuicao_planos começa na linha 30 e vai até a linha 32.
start_row_dist = 30
end_row_dist = 32

data2 = Reference(ws_calc, min_col=2, min_row=start_row_dist - 1, max_row=end_row_dist, max_col=2) # -1 para incluir o cabeçalho
cats2 = Reference(ws_calc, min_col=1, min_row=start_row_dist, max_row=end_row_dist)

chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.height = 10 # Altura em cm
chart2.width = 15 # Largura em cm
ws_dash.add_chart(chart2, "G8") # Posicionar ao lado do primeiro gráfico

# --- 4.3. Salvar o arquivo ---
wb.save('dashboard_vendas_final.xlsx')

print("Dashboard de vendas gerado com sucesso em 'dashboard_vendas_final.xlsx'")
